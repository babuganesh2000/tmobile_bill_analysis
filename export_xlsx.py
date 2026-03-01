"""Export all DuckDB tables to a formatted Excel workbook."""

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DB_PATH = r"c:\tmobile\tmobile_bills.duckdb"
XLSX_PATH = r"c:\tmobile\TMobile_Bills.xlsx"

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="E20074", end_color="E20074", fill_type="solid")  # T-Mobile magenta
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
MONEY_FMT = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'

SHEETS = [
    {
        "name": "Invoices",
        "query": """
            SELECT bill_date, bill_month, total_due, plans_total, equipment_total,
                   services_total, onetime_charges, total_savings, previous_balance,
                   voice_line_count, connected_device_count, wearable_count,
                   data_used_gb, due_date, account_number, file_name
            FROM invoices ORDER BY bill_date
        """,
        "headers": ["Bill Date", "Month", "Total Due", "Plans", "Equipment",
                     "Services", "One-Time", "Total Savings", "Prev Balance",
                     "Voice Lines", "Connected Devices", "Wearables",
                     "Data Used (GB)", "Due Date", "Account", "File"],
        "money_cols": [3,4,5,6,7,8,9,13],
        "date_cols": [1,14],
        "int_cols": [10,11,12],
    },
    {
        "name": "Line Charges",
        "query": """
            SELECT lc.bill_date, lc.phone_number, lc.line_type, lc.status, lc.annotation,
                   lc.plans_charge, lc.equipment_charge, lc.services_charge,
                   lc.onetime_charge, lc.line_total, i.bill_month
            FROM line_charges lc
            JOIN invoices i ON i.file_name = lc.file_name
            ORDER BY lc.bill_date, lc.phone_number
        """,
        "headers": ["Bill Date", "Phone Number", "Line Type", "Status", "Annotation",
                     "Plans", "Equipment", "Services", "One-Time", "Line Total", "Month"],
        "money_cols": [6,7,8,9,10],
        "date_cols": [1],
    },
    {
        "name": "Lines",
        "query": """
            SELECT phone_number, line_type, first_seen_date, last_seen_date,
                   months_active, is_current, lifecycle_notes
            FROM lines ORDER BY line_type, first_seen_date, phone_number
        """,
        "headers": ["Phone Number", "Line Type", "First Seen", "Last Seen",
                     "Months Active", "Current?", "Lifecycle Notes"],
        "date_cols": [3,4],
        "int_cols": [5],
    },
    {
        "name": "Persons",
        "query": """
            SELECT p.person_id, p.person_label, pl.phone_number, pl.relationship,
                   l.line_type, l.first_seen_date, l.last_seen_date,
                   l.months_active, l.is_current
            FROM person_lines pl
            JOIN persons p ON p.person_id = pl.person_id
            JOIN lines l ON l.phone_number = pl.phone_number
            ORDER BY p.person_id
        """,
        "headers": ["Person ID", "Person Label", "Phone Number", "Relationship",
                     "Line Type", "First Seen", "Last Seen", "Months Active", "Current?"],
        "date_cols": [6,7],
        "int_cols": [1,8],
    },
    {
        "name": "Person Monthly Cost",
        "query": """
            SELECT person_id, person_label, bill_date, phone_number, line_type,
                   status, plans_charge, equipment_charge, services_charge,
                   onetime_charge, line_total
            FROM person_monthly_cost
            ORDER BY person_id, bill_date
        """,
        "headers": ["Person ID", "Person", "Bill Date", "Phone", "Line Type",
                     "Status", "Plans", "Equipment", "Services", "One-Time", "Total"],
        "money_cols": [7,8,9,10,11],
        "date_cols": [3],
        "int_cols": [1],
    },
    {
        "name": "Person Totals",
        "query": """
            SELECT person_id, person_label, months,
                   total_plans, total_equipment, total_services,
                   total_onetime, grand_total
            FROM person_total ORDER BY person_id
        """,
        "headers": ["Person ID", "Person", "Months", "Total Plans", "Total Equipment",
                     "Total Services", "Total One-Time", "Grand Total"],
        "money_cols": [4,5,6,7,8],
        "int_cols": [1,3],
    },
    {
        "name": "Savings",
        "query": """
            SELECT s.bill_date, i.bill_month, s.discount_type, s.amount
            FROM savings s
            JOIN invoices i ON i.file_name = s.file_name
            ORDER BY s.bill_date, s.discount_type
        """,
        "headers": ["Bill Date", "Month", "Discount Type", "Amount"],
        "money_cols": [4],
        "date_cols": [1],
    },
    {
        "name": "Payments",
        "query": """
            SELECT p.bill_date, i.bill_month, p.entry_type, p.description, p.amount
            FROM payments p
            JOIN invoices i ON i.file_name = p.file_name
            ORDER BY p.bill_date
        """,
        "headers": ["Bill Date", "Month", "Type", "Description", "Amount"],
        "money_cols": [5],
        "date_cols": [1],
    },
    {
        "name": "Line Usage",
        "query": """
            SELECT bill_date, phone_number, talk_minutes, data_mb, data_gb, file_name
            FROM line_usage
            ORDER BY bill_date, phone_number
        """,
        "headers": ["Bill Date", "Phone Number", "Talk Minutes", "Data (MB)", "Data (GB)", "File"],
        "date_cols": [1],
        "int_cols": [3],
        "decimal_cols": [4, 5],
    },
    {
        "name": "Account Usage",
        "query": """
            SELECT au.bill_date, i.bill_month, au.total_talk_minutes,
                   au.total_messages, au.total_data_gb, i.total_due
            FROM account_usage au
            JOIN invoices i ON i.file_name = au.file_name
            ORDER BY au.bill_date
        """,
        "headers": ["Bill Date", "Month", "Total Talk (min)", "Total Messages",
                     "Total Data (GB)", "Total Due"],
        "date_cols": [1],
        "int_cols": [3, 4],
        "decimal_cols": [5],
        "money_cols": [6],
    },
    {
        "name": "Line Usage Monthly",
        "query": """
            SELECT bill_date, bill_month, phone_number, line_type,
                   talk_minutes, data_mb, data_gb, line_total
            FROM line_usage_monthly
            ORDER BY bill_date, phone_number
        """,
        "headers": ["Bill Date", "Month", "Phone Number", "Line Type",
                     "Talk Minutes", "Data (MB)", "Data (GB)", "Line Total"],
        "date_cols": [1],
        "int_cols": [5],
        "decimal_cols": [6, 7],
        "money_cols": [8],
    },
    {
        "name": "Usage Summary",
        "query": """
            SELECT bill_date, bill_month, total_talk_minutes, total_messages,
                   total_data_gb, total_due, voice_line_count,
                   connected_device_count, wearable_count
            FROM usage_summary
            ORDER BY bill_date
        """,
        "headers": ["Bill Date", "Month", "Total Talk (min)", "Total Messages",
                     "Total Data (GB)", "Total Due", "Voice Lines",
                     "Connected Devices", "Wearables"],
        "date_cols": [1],
        "int_cols": [3, 4, 7, 8, 9],
        "decimal_cols": [5],
        "money_cols": [6],
    },
]


def write_sheet(ws, spec, rows):
    headers = spec["headers"]
    money_cols = set(spec.get("money_cols", []))
    date_cols = set(spec.get("date_cols", []))
    int_cols = set(spec.get("int_cols", []))
    decimal_cols = set(spec.get("decimal_cols", []))

    # Header row
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            # Convert booleans to Y/N
            if isinstance(val, bool):
                cell.value = "Y" if val else "N"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = val

            cell.border = THIN_BORDER

            if c_idx in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif c_idx in decimal_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif c_idx in date_cols:
                cell.number_format = DATE_FMT
            elif c_idx in int_cols:
                cell.alignment = Alignment(horizontal="center")

    # Zebra striping
    stripe_fill = PatternFill(start_color="FFF0F6", end_color="FFF0F6", fill_type="solid")
    for r_idx in range(3, ws.max_row + 1, 2):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = stripe_fill

    # Auto-fit column widths
    for c_idx in range(1, len(headers) + 1):
        max_len = len(headers[c_idx - 1])
        for r_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 35)

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_pivot_sheet(ws, con):
    """Write the per-line monthly cost pivot with account share as a formatted sheet."""
    import pandas as pd

    # Line-level charges
    df = con.execute("""
        SELECT lc.phone_number,
               i.bill_month || ' ' || EXTRACT(YEAR FROM i.bill_date)::VARCHAR AS month_label,
               lc.line_total, i.bill_date
        FROM line_charges lc
        JOIN invoices i ON i.file_name = lc.file_name
        ORDER BY i.bill_date
    """).df()

    line_df = df[df["phone_number"] != "Account"]
    acct_df = df[df["phone_number"] == "Account"]

    months_order = con.execute("""
        SELECT DISTINCT bill_month || ' ' || EXTRACT(YEAR FROM bill_date)::VARCHAR AS ml
        FROM invoices ORDER BY bill_date
    """).df()["ml"].tolist()

    # Per-month account charge & active line count
    acct_per_month = acct_df.groupby("month_label")["line_total"].sum()
    lines_per_month = line_df.groupby("month_label")["phone_number"].nunique()
    share_per_month = (acct_per_month / lines_per_month).fillna(0)

    # Build line charge pivot
    line_pivot = line_df.pivot_table(index="phone_number", columns="month_label",
                                      values="line_total", aggfunc="sum")
    line_pivot = line_pivot[[m for m in months_order if m in line_pivot.columns]]

    # Build account share pivot (same shape)
    acct_share_pivot = line_pivot.copy()
    for m in acct_share_pivot.columns:
        acct_share_pivot[m] = acct_share_pivot[m].apply(
            lambda v: round(share_per_month.get(m, 0), 2) if pd.notna(v) else None
        )

    # Combined = line + share
    combined = line_pivot.fillna(0) + acct_share_pivot.fillna(0)
    combined[line_pivot.isna()] = None

    # Summary columns
    combined["Line TOTAL"] = line_pivot.sum(axis=1, min_count=1)
    combined["Acct Share"] = acct_share_pivot.sum(axis=1, min_count=1)
    combined["Total Due"] = combined["Line TOTAL"].fillna(0) + combined["Acct Share"].fillna(0)

    # TOTAL row
    combined.loc["TOTAL"] = combined.sum()

    # Shorten month labels
    short = {"September":"Sep","October":"Oct","November":"Nov","December":"Dec",
             "January":"Jan","February":"Feb","March":"Mar","April":"Apr",
             "May":"May","June":"Jun","July":"Jul","August":"Aug"}
    def shorten(col):
        if col in ("Line TOTAL", "Acct Share", "Total Due"):
            return col
        parts = col.split()
        return (short.get(parts[0], parts[0][:3]) + " " + parts[1][2:]) if len(parts) == 2 else col
    col_labels = [shorten(c) for c in combined.columns]

    # Header row: "Phone" + month columns + summary columns
    headers = ["Phone"] + col_labels
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    total_fill = PatternFill(start_color="E20074", end_color="E20074", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    stripe_fill = PatternFill(start_color="FFF0F6", end_color="FFF0F6", fill_type="solid")
    acct_share_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    total_due_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    bold_font = Font(name="Calibri", bold=True, size=11)

    # Identify summary column indices (1-based, offset by Phone col)
    summary_cols_map = {}
    for ci, lbl in enumerate(col_labels, 2):
        if lbl == "Line TOTAL":
            summary_cols_map[ci] = "line_total"
        elif lbl == "Acct Share":
            summary_cols_map[ci] = "acct_share"
        elif lbl == "Total Due":
            summary_cols_map[ci] = "total_due"

    for r_idx, (phone, row_data) in enumerate(combined.iterrows(), 2):
        is_total_row = (phone == "TOTAL")
        cell = ws.cell(row=r_idx, column=1, value=phone)
        cell.border = THIN_BORDER
        if is_total_row:
            cell.font = total_font
            cell.fill = total_fill
        elif r_idx % 2 == 0:
            cell.fill = stripe_fill

        for c_idx, val in enumerate(row_data, 2):
            cell = ws.cell(row=r_idx, column=c_idx)
            if pd.isna(val):
                cell.value = None
            else:
                cell.value = round(val, 2)
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
            cell.border = THIN_BORDER
            if is_total_row:
                cell.font = total_font
                cell.fill = total_fill
            elif c_idx in summary_cols_map:
                cell.font = bold_font
                kind = summary_cols_map[c_idx]
                if kind == "acct_share":
                    cell.fill = acct_share_fill
                elif kind == "total_due":
                    cell.fill = total_due_fill
            elif r_idx % 2 == 0:
                cell.fill = stripe_fill

    # Auto-fit column widths
    for c_idx in range(1, len(headers) + 1):
        max_len = len(headers[c_idx - 1])
        for r_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 20)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def main():
    con = duckdb.connect(DB_PATH, read_only=True)
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    for spec in SHEETS:
        print(f"  Writing sheet: {spec['name']}")
        rows = con.execute(spec["query"]).fetchall()
        ws = wb.create_sheet(title=spec["name"])
        write_sheet(ws, spec, rows)
        print(f"    {len(rows)} rows")

    # Pivot sheet: Line Cost by Month
    print(f"  Writing sheet: Line Cost by Month")
    ws_pivot = wb.create_sheet(title="Line Cost by Month")
    write_pivot_sheet(ws_pivot, con)
    print(f"    pivot table written")

    con.close()

    wb.save(XLSX_PATH)
    print(f"\nSaved: {XLSX_PATH}")


if __name__ == "__main__":
    main()
